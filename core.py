#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CatchExcelImage.py
提取Excel文件中的图片，支持嵌入式（DISPIMG）和浮动式图片。
支持四种提取粒度：
    1. 整个工作簿
    2. 指定工作表
    3. 指定工作表某一列
    4. 通过图片ID直接提取
"""

from __future__ import annotations
import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from typing import Dict, List, Tuple, Optional

# ----------------------------------------------------------
# 内部工具函数
# ----------------------------------------------------------
def _extract_dispimg_ids(ws: openpyxl.worksheet.worksheet.Worksheet,
                         target_col: Optional[str] = None) -> List[str]:
    """
    从 openpyxl 工作表对象里提取所有 DISPIMG 的图片 ID。
    如果指定了 target_col（如 'A'），则只扫描该列。
    """
    ids = []
    col_idx = openpyxl.utils.column_index_from_string(target_col) if target_col else None

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if target_col and cell.column != col_idx:
                continue
            value = str(cell.value or "")
            if '=_xlfn.DISPIMG(' in value:
                start = value.find('"') + 1
                end = value.find('"', start)
                ids.append(value[start:end])
    return ids


def _build_id_to_image_map(xlsx_path: str) -> Dict[str, str]:
    """
    解析 .xlsx 内部结构，返回 {image_id -> image内部路径} 的映射
    """
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 检查必要的文件是否存在
        file_list = z.namelist()
        if 'xl/cellimages.xml' not in file_list:
            raise KeyError("No embedded images found: 'xl/cellimages.xml' not found")
        if 'xl/_rels/cellimages.xml.rels' not in file_list:
            raise KeyError("No embedded images found: 'xl/_rels/cellimages.xml.rels' not found")
        
        cellimages_xml = z.read('xl/cellimages.xml')
        rels_xml = z.read('xl/_rels/cellimages.xml.rels')

    root = ET.fromstring(cellimages_xml)
    root_rels = ET.fromstring(rels_xml)

    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/package/2006/relationships'
    }

    # 1. name -> rid
    name_to_rid = {}
    for pic in root.findall('.//xdr:pic', namespaces):
        name = pic.find('.//xdr:cNvPr', namespaces).attrib['name']
        rid = pic.find('.//a:blip', namespaces).attrib[
            '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
        name_to_rid[name] = rid

    # 2. rid -> 内部路径
    rid_to_path = {}
    for rel in root_rels.findall('.//r:Relationship', namespaces):
        rid_to_path[rel.attrib['Id']] = rel.attrib['Target']

    return {name: rid_to_path[rid] for name, rid in name_to_rid.items() if rid in rid_to_path}


def _extract_floating_images(xlsx_path: str) -> Dict[str, str]:
    """
    提取Excel文件中的浮动式图片（非嵌入式图片）
    返回 {图片文件名 -> 内部路径} 的映射
    """
    floating_images = {}
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 获取所有文件列表
        file_list = z.namelist()
        
        # 查找xl/media目录下的图片文件
        for file_path in file_list:
            if file_path.startswith('xl/media/') and not file_path.endswith('/'):
                # 提取文件名（不包含路径）
                filename = file_path.split('/')[-1]
                # 去掉xl/前缀，因为后续处理会加上
                internal_path = file_path[3:]  # 去掉'xl/'前缀
                floating_images[filename] = internal_path
    
    return floating_images


def _get_all_floating_image_positions(xlsx_path: str) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    获取所有工作表中浮动图片的位置信息
    返回 {工作表名 -> {图片文件名 -> {'from_col': int, 'to_col': int, 'from_row': int, 'to_row': int}}} 的映射
    """
    all_positions = {}
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    
    for sheet_index, ws in enumerate(wb.worksheets, 1):
        sheet_name = ws.title
        all_positions[sheet_name] = _get_floating_image_positions(xlsx_path, sheet_name)
    
    return all_positions


def _get_floating_image_positions(xlsx_path: str, sheet_name: str) -> Dict[str, Dict[str, int]]:
    """
    获取指定工作表中浮动图片的位置信息
    返回 {图片文件名 -> {'from_col': int, 'to_col': int, 'from_row': int, 'to_row': int}} 的映射
    """
    image_positions = {}
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        file_list = z.namelist()
        
        # 首先找到工作表对应的drawing文件
        # 需要解析xl/worksheets/_rels/sheet{N}.xml.rels来找到对应的drawing文件
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        sheet_index = None
        for i, ws in enumerate(wb.worksheets, 1):
            if ws.title == sheet_name:
                sheet_index = i
                break
        
        if sheet_index is None:
            return image_positions
        
        # 查找对应的drawing文件
        drawing_rels_path = f'xl/worksheets/_rels/sheet{sheet_index}.xml.rels'
        drawing_path = None
        
        if drawing_rels_path in file_list:
            try:
                rels_xml = z.read(drawing_rels_path)
                rels_root = ET.fromstring(rels_xml)
                
                # 查找drawing关系
                for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    if 'drawing' in rel.get('Target', ''):
                        drawing_path = 'xl/' + rel.get('Target')
                        break
            except:
                pass
        
        if drawing_path and drawing_path in file_list:
            try:
                drawing_xml = z.read(drawing_path)
                drawing_root = ET.fromstring(drawing_xml)
                
                # 定义命名空间
                namespaces = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                # 解析图片位置信息
                for pic in drawing_root.findall('.//xdr:pic', namespaces):
                    try:
                        # 获取图片的关系ID
                        blip = pic.find('.//a:blip', namespaces)
                        if blip is not None:
                            r_embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                            
                            # 获取位置信息
                            anchor = pic.getparent()
                            from_elem = anchor.find('.//xdr:from', namespaces)
                            to_elem = anchor.find('.//xdr:to', namespaces)
                            
                            if from_elem is not None and to_elem is not None:
                                from_col_elem = from_elem.find('xdr:col', namespaces)
                                to_col_elem = to_elem.find('xdr:col', namespaces)
                                from_row_elem = from_elem.find('xdr:row', namespaces)
                                to_row_elem = to_elem.find('xdr:row', namespaces)
                                
                                if all(elem is not None for elem in [from_col_elem, to_col_elem, from_row_elem, to_row_elem]):
                                    # 通过关系ID找到对应的图片文件名
                                    drawing_rels_path = drawing_path.replace('.xml', '.xml.rels')
                                    if drawing_rels_path in file_list:
                                        drawing_rels_xml = z.read(drawing_rels_path)
                                        drawing_rels_root = ET.fromstring(drawing_rels_xml)
                                        
                                        for rel in drawing_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                            if rel.get('Id') == r_embed:
                                                target = rel.get('Target')
                                                if target:
                                                    filename = target.split('/')[-1]
                                                    image_positions[filename] = {
                                                        'from_col': int(from_col_elem.text),
                                                        'to_col': int(to_col_elem.text),
                                                        'from_row': int(from_row_elem.text),
                                                        'to_row': int(to_row_elem.text)
                                                    }
                                                break
                    except:
                        continue
            except:
                pass
    
    return image_positions


def _get_row_data_for_image(ws: openpyxl.worksheet.worksheet.Worksheet, 
                           img_id: str, 
                           target_columns: List[str]) -> Dict[str, any]:
    """
    获取包含指定图片ID的行的数据
    
    Args:
        ws: 工作表对象
        img_id: 图片ID
        target_columns: 需要获取数据的列名列表
    
    Returns:
        包含列名和对应值的字典
    """
    row_data = {}
    
    # 遍历工作表查找包含该图片ID的单元格
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            value = str(cell.value or "")
            if '=_xlfn.DISPIMG(' in value and img_id in value:
                # 找到包含该图片ID的行，获取指定列的数据
                row_num = cell.row
                for col_name in target_columns:
                    try:
                        col_idx = openpyxl.utils.column_index_from_string(col_name)
                        cell_value = ws.cell(row=row_num, column=col_idx).value
                        row_data[col_name] = cell_value
                    except:
                        row_data[col_name] = None
                return row_data
    
    return row_data


def _get_all_images_map(xlsx_path: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    获取Excel文件中所有图片的映射
    返回 (嵌入式图片映射, 浮动式图片映射)
    """
    # 尝试获取嵌入式图片，如果文件不存在则返回空字典
    try:
        embedded_images = _build_id_to_image_map(xlsx_path)
    except KeyError:
        # 如果没有嵌入式图片相关文件，返回空字典
        embedded_images = {}
    
    floating_images = _extract_floating_images(xlsx_path)
    return embedded_images, floating_images


# ----------------------------------------------------------
# 对外 API
# ----------------------------------------------------------
def extract_workbook_images(xlsx_path: str,
                            output_dir: str = 'images',
                            include_floating: bool = True,
                            custom_naming_func=None) -> List[str]:
    """
    提取整个工作簿里所有图片（包括嵌入式DISPIMG图片和浮动式图片）。
    
    Args:
        xlsx_path: Excel文件路径
        output_dir: 输出目录
        include_floating: 是否包含浮动式图片，默认True
    
    Returns:
        已保存图片的绝对路径列表
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    all_ids = []
    id_to_worksheet = {}  # 记录每个图片ID对应的工作表
    
    for ws in wb.worksheets:
        ws_ids = _extract_dispimg_ids(ws)
        all_ids.extend(ws_ids)
        # 记录每个图片ID对应的工作表
        for img_id in ws_ids:
            id_to_worksheet[img_id] = ws

    embedded_images, floating_images = _get_all_images_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    sequence_counter = 1
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 提取嵌入式图片
        for img_id in set(all_ids):
            if img_id in embedded_images:
                img_internal = 'xl/' + embedded_images[img_id]
                img_data = z.read(img_internal)
                
                # 生成文件名
                if custom_naming_func:
                    # 获取该图片所在的工作表和行数据
                    ws = id_to_worksheet.get(img_id)
                    row_data = None
                    if ws:
                        # 获取所有列的数据（A-Z）
                        all_columns = [openpyxl.utils.get_column_letter(i) for i in range(1, 27)]
                        row_data = _get_row_data_for_image(ws, img_id, all_columns)
                    filename = custom_naming_func(img_id, row_data, sequence_counter)
                    out_file = os.path.join(output_dir, f"{filename}.png")
                else:
                    out_file = os.path.join(output_dir, f"{img_id}.png")
                
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
                sequence_counter += 1
        
        # 提取浮动式图片
        if include_floating:
            for filename, internal_path in floating_images.items():
                img_internal = 'xl/' + internal_path
                img_data = z.read(img_internal)
                # 保持原始文件扩展名
                file_ext = os.path.splitext(filename)[1] or '.png'
                
                # 生成文件名
                if custom_naming_func:
                    base_name = os.path.splitext(filename)[0]
                    custom_filename = custom_naming_func(base_name, None, sequence_counter)
                    out_file = os.path.join(output_dir, f"{custom_filename}{file_ext}")
                else:
                    out_file = os.path.join(output_dir, f"FLOAT_{filename}")
                
                with open(out_file, 'wb') as f:
                     f.write(img_data)
                saved.append(os.path.abspath(out_file))
                sequence_counter += 1
    
    return saved


def extract_sheet_images(xlsx_path: str,
                         sheet_name: str,
                         output_dir: str = 'images',
                         include_floating: bool = True,
                         custom_naming_func=None) -> List[str]:
    """
    提取指定工作表中的图片（包括嵌入式DISPIMG图片和浮动式图片）。
    
    Args:
        xlsx_path: Excel文件路径
        sheet_name: 工作表名称
        output_dir: 输出目录
        include_floating: 是否包含浮动式图片，默认True
    
    Returns:
        已保存图片的绝对路径列表
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    ids = _extract_dispimg_ids(ws)

    embedded_images, floating_images = _get_all_images_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    sequence_counter = 1
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 提取嵌入式图片
        for img_id in set(ids):
            if img_id in embedded_images:
                img_internal = 'xl/' + embedded_images[img_id]
                img_data = z.read(img_internal)
                
                # 生成文件名
                if custom_naming_func:
                    # 获取所有列的数据（A-Z）
                    all_columns = [openpyxl.utils.get_column_letter(i) for i in range(1, 27)]
                    row_data = _get_row_data_for_image(ws, img_id, all_columns)
                    filename = custom_naming_func(img_id, row_data, sequence_counter)
                    out_file = os.path.join(output_dir, f"{filename}.png")
                else:
                    out_file = os.path.join(output_dir, f"{img_id}.png")
                
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
                sequence_counter += 1
        
        # 提取浮动式图片（只提取指定工作表中的浮动图片）
        if include_floating:
            # 获取浮动图片的位置信息
            floating_positions = _get_floating_image_positions(xlsx_path, sheet_name)
            
            for filename, internal_path in floating_images.items():
                 # 检查图片是否在指定工作表中
                 if filename in floating_positions:
                     print(f"浮动图片 {filename} 位于工作表 '{sheet_name}' 中，将被包含")
                     img_internal = 'xl/' + internal_path
                     img_data = z.read(img_internal)
                     # 保持原始文件扩展名
                     file_ext = os.path.splitext(filename)[1] or '.png'
                     
                     # 生成文件名
                     if custom_naming_func:
                         base_name = os.path.splitext(filename)[0]
                         custom_filename = custom_naming_func(base_name, None, sequence_counter)
                         out_file = os.path.join(output_dir, f"{custom_filename}{file_ext}")
                     else:
                         out_file = os.path.join(output_dir, f"FLOAT_{filename}")
                     
                     with open(out_file, 'wb') as f:
                         f.write(img_data)
                     saved.append(os.path.abspath(out_file))
                     sequence_counter += 1
    
    return saved


def extract_column_images(xlsx_path: str,
                          sheet_name: str,
                          columns: str,
                          output_dir: str = 'images',
                          include_floating: bool = True,
                          custom_naming_func=None) -> List[str]:
    """
    提取指定工作表某些列中的图片（包括嵌入式DISPIMG图片和浮动式图片）。
    
    Args:
        xlsx_path: Excel文件路径
        sheet_name: 工作表名称
        columns: 列名，支持单列(如'A')或多列(如'A,B,C'或'A-C')
        output_dir: 输出目录
        include_floating: 是否包含浮动式图片，默认True
    
    Returns:
        已保存图片的绝对路径列表
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    
    # 解析列参数
    column_list = []
    if ',' in columns:
        # 逗号分隔的多列: A,B,C
        column_list = [col.strip() for col in columns.split(',')]
    elif '-' in columns:
        # 范围表示: A-C
        start_col, end_col = columns.split('-')
        start_idx = openpyxl.utils.column_index_from_string(start_col.strip())
        end_idx = openpyxl.utils.column_index_from_string(end_col.strip())
        for i in range(start_idx, end_idx + 1):
            column_list.append(openpyxl.utils.get_column_letter(i))
    else:
        # 单列
        column_list = [columns.strip()]
    
    # 提取所有指定列的图片ID
    all_ids = []
    for col in column_list:
        ids = _extract_dispimg_ids(ws, target_col=col)
        all_ids.extend(ids)

    embedded_images, floating_images = _get_all_images_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    sequence_counter = 1
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 提取嵌入式图片
        for img_id in set(all_ids):
            if img_id in embedded_images:
                img_internal = 'xl/' + embedded_images[img_id]
                img_data = z.read(img_internal)
                
                # 生成文件名
                if custom_naming_func:
                    # 对于Excel列命名模式，需要获取行数据
                    row_data = _get_row_data_for_image(ws, img_id, column_list)
                    filename = custom_naming_func(img_id, row_data, sequence_counter)
                    out_file = os.path.join(output_dir, f"{filename}.png")
                else:
                    out_file = os.path.join(output_dir, f"{img_id}.png")
                
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
                sequence_counter += 1
        
        # 提取浮动式图片（根据指定列进行过滤）
        if include_floating:
            # 获取浮动图片的位置信息
            floating_positions = _get_floating_image_positions(xlsx_path, sheet_name)
            
            # 将列名转换为列索引（0-based）
            target_col_indices = set()
            for col in column_list:
                col_idx = openpyxl.utils.column_index_from_string(col) - 1  # 转换为0-based索引
                target_col_indices.add(col_idx)
            
            for filename, internal_path in floating_images.items():
                 # 检查图片是否在指定列范围内
                 should_include = False
                 if filename in floating_positions:
                     pos = floating_positions[filename]
                     # 检查图片的列范围是否与指定列有交集
                     img_col_range = set(range(pos['from_col'], pos['to_col'] + 1))
                     if img_col_range.intersection(target_col_indices):
                         should_include = True
                         print(f"浮动图片 {filename} 位于列 {pos['from_col']}-{pos['to_col']}，与指定列 {columns} 有交集，将被包含")
                     else:
                         print(f"浮动图片 {filename} 位于列 {pos['from_col']}-{pos['to_col']}，与指定列 {columns} 无交集，已过滤")
                 else:
                     # 如果无法获取位置信息，为了保持兼容性，仍然包含该图片
                     print(f"警告：无法获取浮动图片 {filename} 的位置信息，为保持兼容性仍然包含")
                     should_include = True
                 
                 if should_include:
                     img_internal = 'xl/' + internal_path
                     img_data = z.read(img_internal)
                     # 保持原始文件扩展名
                     file_ext = os.path.splitext(filename)[1] or '.png'
                     
                     # 生成文件名
                     if custom_naming_func:
                         base_name = os.path.splitext(filename)[0]
                         custom_filename = custom_naming_func(base_name, None, sequence_counter)
                         out_file = os.path.join(output_dir, f"{custom_filename}{file_ext}")
                     else:
                         out_file = os.path.join(output_dir, f"FLOAT_{filename}")
                     
                     with open(out_file, 'wb') as f:
                         f.write(img_data)
                     saved.append(os.path.abspath(out_file))
                     sequence_counter += 1
    
    return saved


def extract_image_by_id(xlsx_path: str,
                        image_id: str,
                        output_dir: str = 'images',
                        custom_naming_func=None) -> Optional[str]:
    """
    通过指定的图片ID直接提取嵌入式图片。
    只支持嵌入式图片ID。
    
    Args:
        xlsx_path: Excel文件路径
        image_id: 嵌入式图片ID
        output_dir: 输出目录
    
    Returns:
        保存的图片绝对路径，如果图片不存在则返回None
    """
    try:
        embedded_images, _ = _get_all_images_map(xlsx_path)
    except KeyError:
        # 如果没有嵌入式图片相关文件，返回None
        return None
    
    os.makedirs(output_dir, exist_ok=True)
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # 只处理嵌入式图片ID
        if image_id in embedded_images:
            img_internal = 'xl/' + embedded_images[image_id]
            img_data = z.read(img_internal)
            
            # 生成文件名
            if custom_naming_func:
                # 查找包含该图片ID的工作表和行数据
                wb = openpyxl.load_workbook(xlsx_path, data_only=False)
                row_data = None
                for ws in wb.worksheets:
                    # 获取所有列的数据（A-Z）
                    all_columns = [openpyxl.utils.get_column_letter(i) for i in range(1, 27)]
                    temp_row_data = _get_row_data_for_image(ws, image_id, all_columns)
                    if temp_row_data:  # 如果找到了数据，说明图片在这个工作表中
                        row_data = temp_row_data
                        break
                filename = custom_naming_func(image_id, row_data, 1)
                out_file = os.path.join(output_dir, f"{filename}.png")
            else:
                # 修复文件名重复ID问题：直接使用image_id作为文件名，不再添加ID_前缀
                out_file = os.path.join(output_dir, f"{image_id}.png")
            
            with open(out_file, 'wb') as f:
                f.write(img_data)
            return os.path.abspath(out_file)
    
    return None





def get_embedded_image_ids(xlsx_path: str) -> List[str]:
    """
    仅获取嵌入式图片ID列表。
    
    Returns:
        嵌入式图片ID的列表
    """
    embedded_images, _ = _get_all_images_map(xlsx_path)
    return list(embedded_images.keys())


def get_floating_image_names(xlsx_path: str) -> List[str]:
    """
    仅获取浮动式图片文件名列表。
    
    Returns:
        浮动式图片文件名的列表
    """
    _, floating_images = _get_all_images_map(xlsx_path)
    return list(floating_images.keys())
